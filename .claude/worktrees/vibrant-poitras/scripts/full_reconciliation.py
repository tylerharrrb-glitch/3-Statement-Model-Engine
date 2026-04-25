"""
FULL RECONCILIATION: Compare Excel formulas vs Engine-computed values
Reads the WOLF Excel file and compares the Scenarios sheet (engine data)
against the IS/BS/CF formulas to find discrepancies.
"""
import openpyxl
from openpyxl.utils import get_column_letter
import os, sys, json

sys.stdout.reconfigure(encoding='utf-8')

SRC = os.path.join(os.environ['USERPROFILE'], 'Downloads', 'WOLF_Financial_Model.xlsx')

# Load with formulas
wb = openpyxl.load_workbook(SRC)
# Load with data (cached values)
wb_data = openpyxl.load_workbook(SRC, data_only=True)

print("=" * 100)
print("FULL EXCEL RECONCILIATION ANALYSIS")
print("=" * 100)

# ══════════════════════════════════════════════════════════
# 1. MAP ALL SHEETS: rows, labels, formulas
# ══════════════════════════════════════════════════════════

def map_sheet(sheet_name, wb_formula, wb_data):
    """Map a sheet: for each row with a label in col A, show the formula and cached value for cols B-H."""
    ws = wb_formula[sheet_name]
    ws_d = wb_data[sheet_name]
    rows = []
    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if not label or not str(label).strip():
            continue
        label = str(label).strip()
        row_data = {'row': r, 'label': label, 'formulas': {}, 'values': {}}
        for c in range(2, ws.max_column + 1):
            col_letter = get_column_letter(c)
            formula = ws.cell(r, c).value
            value = ws_d.cell(r, c).value
            if formula is not None or value is not None:
                row_data['formulas'][col_letter] = str(formula)[:80] if formula else None
                row_data['values'][col_letter] = value
        rows.append(row_data)
    return rows

# Map Cash Flow Statement
print("\n" + "=" * 100)
print("CASH FLOW STATEMENT — Formula Audit")
print("=" * 100)

cf_rows = map_sheet('Cash Flow Statement', wb, wb_data)
for row in cf_rows:
    label = row['label'][:40]
    # Show formulas for the last projection column (2028E, typically col H)
    # Find the column with 2028E header
    print(f"\n  Row {row['row']:3d}: {label}")
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        f = row['formulas'].get(col)
        v = row['values'].get(col)
        if f or v is not None:
            v_str = f"{v:,.0f}" if isinstance(v, (int, float)) else str(v)
            f_str = str(f)[:60] if f else 'NONE'
            print(f"    {col}: val={v_str:>12s}  formula={f_str}")

# ══════════════════════════════════════════════════════════
# 2. MAP THE SCENARIOS SHEET — Optimistic block
# ══════════════════════════════════════════════════════════
print("\n" + "=" * 100)
print("SCENARIOS SHEET — Optimistic Block Engine-Computed Values")
print("=" * 100)

scen = wb_data['Scenarios']
# Find optimistic block
opt_start = None
for r in range(1, scen.max_row + 1):
    v = str(scen.cell(r, 1).value or '')
    if 'OPTIMISTIC' in v.upper():
        opt_start = r
        break

if opt_start:
    print(f"  Optimistic block starts at row {opt_start}")
    for r in range(opt_start, opt_start + 65):
        label = scen.cell(r, 1).value
        if not label: continue
        vals = []
        for c in range(2, 10):
            v = scen.cell(r, c).value
            if isinstance(v, (int, float)):
                if abs(v) < 2 and abs(v) > 0:
                    vals.append(f"{v*100:.2f}%")
                else:
                    vals.append(f"{v:>12,.0f}")
            elif v:
                vals.append(str(v)[:12])
        if vals:
            print(f"  Row {r:3d}: {str(label)[:40]:40s} → {', '.join(vals[:7])}")

# ══════════════════════════════════════════════════════════
# 3. ASSUMPTIONS SHEET — rows that have IF formulas
# ══════════════════════════════════════════════════════════
print("\n" + "=" * 100)
print("ASSUMPTIONS SHEET — Engine-Computed IF Formula Rows")
print("=" * 100)

asheet = wb['Assumptions']
asheet_d = wb_data['Assumptions']

computed_rows = []
for r in range(1, asheet.max_row + 1):
    label = asheet.cell(r, 1).value
    if not label: continue
    
    # Check if any cell in row has IF formula
    has_if = False
    for c in range(5, 10):
        v = asheet.cell(r, c).value
        if v and isinstance(v, str) and 'IF' in v:
            has_if = True
            break
    
    if has_if:
        computed_rows.append(r)
        # Show what the formula references
        formula = asheet.cell(r, 5).value  # Column E
        base_val = asheet_d.cell(r, 5).value
        print(f"  Row {r:3d}: {str(label)[:40]:40s}  E={base_val}  formula={str(formula)[:80]}")

print(f"\n  Total rows with IF formulas: {len(computed_rows)}")

# ══════════════════════════════════════════════════════════
# 4. INCOME STATEMENT — check which cells are formulas vs hardcoded
# ══════════════════════════════════════════════════════════
print("\n" + "=" * 100)
print("INCOME STATEMENT — Formula vs Hardcoded Audit (Projection only)")
print("=" * 100)

is_ws = wb['Income Statement']
is_ws_d = wb_data['Income Statement']

for r in range(1, is_ws.max_row + 1):
    label = is_ws.cell(r, 1).value
    if not label: continue
    label = str(label).strip()
    
    for c in range(5, 10):  # Projection columns E-I
        val = is_ws.cell(r, c).value
        data_val = is_ws_d.cell(r, c).value
        col = get_column_letter(c)
        
        if val is None: continue
        
        is_formula = isinstance(val, str) and (val.startswith('=') or '+' in str(val) or '-' in str(val))
        is_number = isinstance(val, (int, float))
        
        if is_number:
            print(f"  HARDCODED Row {r}: {label[:35]:35s}  {col}={val:>12,.0f}")
        elif is_formula and c == 5:  # Only print formula for col E
            print(f"  FORMULA  Row {r}: {label[:35]:35s}  E={str(val)[:80]}")

# ══════════════════════════════════════════════════════════
# 5. BALANCE SHEET — same audit
# ══════════════════════════════════════════════════════════
print("\n" + "=" * 100)
print("BALANCE SHEET — Hardcoded Check (Projection only)")
print("=" * 100)

bs_ws = wb['Balance Sheet']
for r in range(1, bs_ws.max_row + 1):
    label = bs_ws.cell(r, 1).value
    if not label: continue
    label = str(label).strip()
    
    for c in range(5, 10):
        val = bs_ws.cell(r, c).value
        if val is None: continue
        is_number = isinstance(val, (int, float))
        
        if is_number:
            col = get_column_letter(c)
            print(f"  HARDCODED Row {r}: {label[:35]:35s}  {col}={val:>12,.0f}")

# ══════════════════════════════════════════════════════════
# 6. CASH FLOW — same audit
# ══════════════════════════════════════════════════════════
print("\n" + "=" * 100)
print("CASH FLOW — Hardcoded Check (Projection only)")
print("=" * 100)

cf_ws = wb['Cash Flow Statement']
for r in range(1, cf_ws.max_row + 1):
    label = cf_ws.cell(r, 1).value
    if not label: continue
    label = str(label).strip()
    
    for c in range(5, 10):
        val = cf_ws.cell(r, c).value
        if val is None: continue
        is_number = isinstance(val, (int, float))
        
        if is_number:
            col = get_column_letter(c)
            print(f"  HARDCODED Row {r}: {label[:35]:35s}  {col}={val:>12,.0f}")

wb.close()
wb_data.close()
print("\n=== Reconciliation Complete ===")
