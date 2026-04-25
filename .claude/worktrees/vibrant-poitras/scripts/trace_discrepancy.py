"""
CORRECTED TRACE: Compute Optimistic Cash from both Engine stored values 
and simulated Excel formulas, using correct base revenue = 950,000 (2023A).
"""
import openpyxl
from openpyxl.utils import get_column_letter
import os, sys

sys.stdout.reconfigure(encoding='utf-8')

SRC = os.path.join(os.environ['USERPROFILE'], 'Downloads', 'WOLF_Financial_Model.xlsx')
wb_d = openpyxl.load_workbook(SRC, data_only=True)

# Get Scenarios Optimistic data
scen = wb_d['Scenarios']
opt_start = None
for r in range(1, scen.max_row + 1):
    if 'OPTIMISTIC' in str(scen.cell(r, 1).value or '').upper():
        opt_start = r
        break

opt = {}
for r in range(opt_start, opt_start + 62):
    label = scen.cell(r, 1).value
    if not label: continue
    label = str(label).strip()
    vals = [scen.cell(r, c).value for c in range(2, 10)]  # B=0..I=7
    opt[label] = vals

# Get historical values from BS/IS (cached = Base Case values but historical is same for all)
bs = wb_d['Balance Sheet']
is_ws = wb_d['Income Statement']

# Historical columns: B=col2=2021A, C=col3=2022A, D=col4=2023A
# 2023A is column D (idx 4 in Excel, idx 2 in 0-based from col B)
rev_2023 = is_ws.cell(2, 4).value  # D2
print(f"2023A Revenue (D2) = {rev_2023:,.0f}")

# BS historical values for 2023A (col D)
cash_2023 = bs.cell(3, 4).value
ar_2023 = bs.cell(4, 4).value
inv_2023 = bs.cell(5, 4).value
prepaid_2023 = bs.cell(6, 4).value
ap_2023 = bs.cell(18, 4).value
accrued_2023 = bs.cell(19, 4).value  
deferred_2023 = bs.cell(22, 4).value

print(f"2023A: Cash={cash_2023:,.0f} AR={ar_2023:,.0f} Inv={inv_2023:,.0f} Prepaid={prepaid_2023:,.0f}")
print(f"       AP={ap_2023:,.0f} Accrued={accrued_2023:,.0f} DeferredRev={deferred_2023:,.0f}")

# Helper: projection index to Scenarios index
# Scenarios has B=idx0..I=idx7. Projection starts at E=idx3
def sp(year):
    """year 0=2024E..4=2028E → Scenarios index 3..7"""
    return 3 + year

def get_opt(key, year):
    idx = sp(year)
    vals = opt.get(key, [])
    return vals[idx] if idx < len(vals) else 0

# ═══════════════════════════════════════════════
# PART 1: Simulate what Excel formulas would produce for Optimistic
# ═══════════════════════════════════════════════
print("\n" + "=" * 100)
print("PART 1: SIMULATED EXCEL (Optimistic)")
print("=" * 100)

# Revenue chain
rev = [0] * 5
prior_rev = rev_2023  # 950,000 (2023A, col D)
for yr in range(5):
    g = get_opt('Revenue Growth Rate', yr)
    rev[yr] = prior_rev * (1 + g)
    prior_rev = rev[yr]

print(f"\nRevenue: {[f'{r:,.0f}' for r in rev]}")

# COGS, SGA, RD, Other OpEx
cogs = [rev[yr] * get_opt('COGS % of Revenue', yr) for yr in range(5)]
gp = [rev[yr] - cogs[yr] for yr in range(5)]
sga = [rev[yr] * get_opt('SG&A % of Revenue', yr) for yr in range(5)]
rd = [rev[yr] * get_opt('R&D % of Revenue', yr) for yr in range(5)]
other_opex = [rev[yr] * get_opt('Other OpEx % of Revenue', yr) for yr in range(5)]
sbc = [get_opt('Stock-Based Comp Amount', yr) for yr in range(5)]
dep = [get_opt('Depreciation (Computed)', yr) for yr in range(5)]  # engine-computed
amort = [get_opt('Amortization Amount', yr) for yr in range(5)]

total_opex = [sga[yr] + rd[yr] + dep[yr] + amort[yr] + other_opex[yr] + sbc[yr] for yr in range(5)]
ebit = [gp[yr] - total_opex[yr] for yr in range(5)]

# Interest (engine-computed)
int_inc = [get_opt('Interest Income (Computed)', yr) for yr in range(5)]
int_exp = [get_opt('Interest Expense (Computed)', yr) for yr in range(5)]

# EBT, Tax, NI
ebt = [ebit[yr] + int_inc[yr] - int_exp[yr] for yr in range(5)]
tax_rate = [get_opt('Tax Rate', yr) for yr in range(5)]
tax = [max(0, ebt[yr] * tax_rate[yr]) for yr in range(5)]
ni = [ebt[yr] - tax[yr] for yr in range(5)]
print(f"Net Income: {[f'{n:,.0f}' for n in ni]}")

# Dividends (Excel formula: -MAX(0, NI * payoutRatio))
div_payout = [get_opt('Dividend Payout Ratio', yr) for yr in range(5)]
dividends_excel = [-max(0, ni[yr] * div_payout[yr]) for yr in range(5)]
dividends_engine = [get_opt('Dividends Paid (Computed)', yr) for yr in range(5)]
print(f"Dividends Excel:  {[f'{d:,.0f}' for d in dividends_excel]}")
print(f"Dividends Engine: {[f'{d:,.0f}' for d in dividends_engine]}")

# WC items
dso = [get_opt('DSO (Days)', yr) for yr in range(5)]
dio = [get_opt('DIO (Days)', yr) for yr in range(5)]
dpo = [get_opt('DPO (Days)', yr) for yr in range(5)]
prepaid_pct = [get_opt('Prepaid % of Revenue', yr) for yr in range(5)]
accrued_pct = [get_opt('Accrued Exp % of Revenue', yr) for yr in range(5)]
deferred_pct = [get_opt('Deferred Rev % of Revenue', yr) for yr in range(5)]

ar = [rev[yr] * dso[yr] / 365 for yr in range(5)]
inv = [cogs[yr] * dio[yr] / 365 for yr in range(5)]
prepaid = [rev[yr] * prepaid_pct[yr] for yr in range(5)]
ap = [cogs[yr] * dpo[yr] / 365 for yr in range(5)]
accrued = [rev[yr] * accrued_pct[yr] for yr in range(5)]
deferred = [rev[yr] * deferred_pct[yr] for yr in range(5)]

# BS items (engine-computed)
ppe = [get_opt('Net PP&E (Computed)', yr) for yr in range(5)]
intangibles = [get_opt('Intangibles (Computed)', yr) for yr in range(5)]
goodwill = [get_opt('Goodwill', yr) for yr in range(5)]
olta = [get_opt('Other Long-Term Assets', yr) for yr in range(5)]
oca = [get_opt('Other Current Assets', yr) for yr in range(5)]
st_debt = [get_opt('Short-Term Debt', yr) for yr in range(5)]
cpltd = [get_opt('Current Portion LTD', yr) for yr in range(5)]
ocl = [get_opt('Other Current Liabilities', yr) for yr in range(5)]
ltd = [get_opt('Long-Term Debt (Computed)', yr) for yr in range(5)]
dtl = [get_opt('Deferred Tax Liabilities', yr) for yr in range(5)]
oltl = [get_opt('Other LT Liabilities', yr) for yr in range(5)]
cs = [get_opt('Common Stock', yr) for yr in range(5)]
apic = [get_opt('APIC (Computed)', yr) for yr in range(5)]
re = [get_opt('Retained Earnings (Computed)', yr) for yr in range(5)]
ts = [get_opt('Treasury Stock (Computed)', yr) for yr in range(5)]
oci = [get_opt('Other Comprehensive Income', yr) for yr in range(5)]

# BS Cash = Total L&E - AR - Inv - Prepaid - OCA - NCA
for yr in range(5):
    total_cur_liab = ap[yr] + accrued[yr] + st_debt[yr] + cpltd[yr] + deferred[yr] + ocl[yr]
    total_non_cur_liab = ltd[yr] + dtl[yr] + oltl[yr]
    total_liab = total_cur_liab + total_non_cur_liab
    total_eq = cs[yr] + apic[yr] + re[yr] + ts[yr] + oci[yr]
    total_le = total_liab + total_eq
    
    nca = ppe[yr] + intangibles[yr] + goodwill[yr] + olta[yr]
    
    # Cash plug:
    cash_plug = total_le - ar[yr] - inv[yr] - prepaid[yr] - oca[yr] - nca
    
    year_label = f"20{24+yr}E"
    print(f"\n{year_label}:")
    print(f"  Total Liabilities = {total_liab:,.0f}")
    print(f"  Total Equity = {total_eq:,.0f} (RE={re[yr]:,.0f})")
    print(f"  Total L+E = {total_le:,.0f}")
    print(f"  AR={ar[yr]:,.0f}  Inv={inv[yr]:,.0f}  Prepaid={prepaid[yr]:,.0f}  OCA={oca[yr]:,.0f}")
    print(f"  NCA={nca:,.0f}")
    print(f"  BS CASH (plug) = {cash_plug:,.0f}")

# Now compute CF Ending Cash
print(f"\n{'='*100}")
print("CF ENDING CASH TRACE")
print(f"{'='*100}")

beg_cash = cash_2023  # 2023A Cash
for yr in range(5):
    # WC changes
    if yr == 0:
        p_ar, p_inv, p_prepaid = ar_2023, inv_2023, prepaid_2023
        p_ap, p_accrued, p_deferred = ap_2023, accrued_2023, deferred_2023
    else:
        p_ar, p_inv, p_prepaid = ar[yr-1], inv[yr-1], prepaid[yr-1]
        p_ap, p_accrued, p_deferred = ap[yr-1], accrued[yr-1], deferred[yr-1]
    
    d_ar = -(ar[yr] - p_ar)
    d_inv = -(inv[yr] - p_inv)
    d_prepaid = -(prepaid[yr] - p_prepaid)
    d_ap = ap[yr] - p_ap
    d_accrued = accrued[yr] - p_accrued
    d_deferred = deferred[yr] - p_deferred
    wc_change = d_ar + d_inv + d_prepaid + d_ap + d_accrued + d_deferred
    
    deferred_tax_change = 0  # DTL doesn't change
    
    cfo = ni[yr] + dep[yr] + amort[yr] + sbc[yr] + deferred_tax_change + wc_change
    
    capex_pct = get_opt('CapEx % of Revenue', yr)
    capex = -abs(rev[yr] * capex_pct)
    cfi = capex + 0 + 0  # acquisitions + asset sales
    
    debt_iss = get_opt('LT Debt Issuance', yr)
    debt_rep = -abs(get_opt('LT Debt Repayment', yr))
    div_paid = dividends_excel[yr]  # Excel formula version
    eq_iss = get_opt('Equity Issuance', yr)
    share_rep = -abs(get_opt('Share Repurchase Amount', yr))
    cff = debt_iss + debt_rep + div_paid + eq_iss + share_rep
    
    net_change = cfo + cfi + cff
    ending_cash = beg_cash + net_change
    
    year_label = f"20{24+yr}E"
    print(f"\n{year_label}:")
    print(f"  NI={ni[yr]:,.0f}  Dep={dep[yr]:,.0f}  Amort={amort[yr]:,.0f}  SBC={sbc[yr]:,.0f}")
    print(f"  WC Change={wc_change:,.0f}  DeferredTax={deferred_tax_change:,.0f}")
    print(f"  CFO={cfo:,.0f}")
    print(f"  CapEx={capex:,.0f}  CFI={cfi:,.0f}")
    print(f"  DebtIss={debt_iss:,.0f}  DebtRep={debt_rep:,.0f}  Div={div_paid:,.0f}")
    print(f"  CFF={cff:,.0f}")
    print(f"  Net Change={net_change:,.0f}")
    print(f"  Beg Cash={beg_cash:,.0f}  End Cash (CF)={ending_cash:,.0f}")
    
    # Now compute BS cash plug for this year (already done above but recompute)
    total_cur_liab = ap[yr] + accrued[yr] + st_debt[yr] + cpltd[yr] + deferred[yr] + ocl[yr]
    total_non_cur_liab = ltd[yr] + dtl[yr] + oltl[yr]
    total_eq = cs[yr] + apic[yr] + re[yr] + ts[yr] + oci[yr]
    total_le = total_cur_liab + total_non_cur_liab + total_eq
    nca = ppe[yr] + intangibles[yr] + goodwill[yr] + olta[yr]
    cash_plug = total_le - ar[yr] - inv[yr] - prepaid[yr] - oca[yr] - nca
    
    print(f"  BS Cash (plug)={cash_plug:,.0f}")
    diff = ending_cash - cash_plug
    print(f"  MISMATCH CF vs BS = {diff:,.0f} {'⚠️' if abs(diff) > 0.01 else '✅'}")
    
    # Use BS plug as next year's beginning cash (like Excel does)
    beg_cash = cash_plug

print(f"\n{'='*100}")
print("PROBLEM: CF Beginning Cash uses BS Cash plug from prior year")
print("But BS Cash is a PLUG that depends on engine-computed RE.")
print("If CF Ending Cash != BS Cash plug, there's a circular mismatch.")
print(f"{'='*100}")

# The KEY question: does the engine's Net Income for Optimistic match
# what the IS formulas would compute?
# Engine RE change = NI - Dividends = NI * 0.95
# So Engine NI = (RE[yr] - RE[yr-1]) / 0.95

re_2023 = opt['Retained Earnings (Computed)'][2]  # col D = 2023A
print(f"\n2023A RE = {re_2023:,.0f}")

print("\nEngine NI (back-computed from RE changes):")
prev_re = re_2023
for yr in range(5):
    re_change = re[yr] - prev_re
    engine_ni = re_change / (1 - div_payout[yr])  # NI = RE_change / (1 - payoutRatio) since RE = RE_prev + NI - Div, and Div = NI * payoutRatio
    
    year_label = f"20{24+yr}E"
    print(f"  {year_label}: RE={re[yr]:,.0f}  RE_change={re_change:,.0f}  Engine_NI={engine_ni:,.0f}  Excel_NI={ni[yr]:,.0f}  DIFF={engine_ni - ni[yr]:,.0f}")
    prev_re = re[yr]

wb_d.close()
