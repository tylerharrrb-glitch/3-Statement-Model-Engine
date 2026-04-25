"""
Phase 3 - Simulate scenario switching by evaluating IF formulas manually.
Since we can't use Excel's recalculation engine, we manually trace what each
IF formula SHOULD evaluate to for each scenario selection.
"""
import openpyxl
import os, sys

sys.stdout.reconfigure(encoding='utf-8')

SRC = os.path.join(os.environ['USERPROFILE'], 'Downloads', 'WOLF_Financial_Model_Fixed.xlsx')

# Load with formulas
wb = openpyxl.load_workbook(SRC)
scen_sheet = wb['Scenarios']
asheet = wb['Assumptions']

# Read all Scenarios data for the 3 blocks
# From diagnosis: Base=rows 7-65, Opt=rows 70-78, Cons=rows 82-90

# Key rows mapping (assumption label -> scenario rows)
KEY_MAPPING = {
    'Revenue Growth Rate': {'assum': 5, 'base': 9, 'opt': 70, 'cons': 82},
    'COGS %':             {'assum': 6, 'base': 10, 'opt': 71, 'cons': 83},
    'SG&A %':             {'assum': 7, 'base': 11, 'opt': 72, 'cons': 84},
    'R&D %':              {'assum': 8, 'base': 12, 'opt': 73, 'cons': 85},
    'Other OpEx %':       {'assum': 9, 'base': 13, 'opt': 74, 'cons': 86},
    'Tax Rate':           {'assum': 10, 'base': 14, 'opt': 75, 'cons': 87},
    'CapEx %':            {'assum': 22, 'base': 26, 'opt': 76, 'cons': 88},
    'Interest Rate':      {'assum': 26, 'base': 30, 'opt': 77, 'cons': 89},
    'Dividend Payout':    {'assum': 32, 'base': 36, 'opt': 78, 'cons': 90},
}

# Load data-only version to read actual values in Scenarios sheet
wb_data = openpyxl.load_workbook(SRC, data_only=True)
scen_data = wb_data['Scenarios']

# Read actual values from Scenarios sheet for each block
print("=" * 100)
print("SCENARIO DATA VALUES (from Scenarios sheet, col E = 2024E)")
print("=" * 100)
print(f"\n{'Metric':<25s} {'Base (row)':>12s} {'Opt (row)':>12s} {'Cons (row)':>12s}")
print("-" * 100)

scenario_values = {}  # (scenario, metric, col) -> value

for metric, rows in KEY_MAPPING.items():
    for scenario, key in [('Base Case', 'base'), ('Optimistic', 'opt'), ('Conservative', 'cons')]:
        for c in range(5, 10):
            v = scen_data.cell(rows[key], c).value
            scenario_values[(scenario, metric, c)] = v

    base_e = scenario_values[('Base Case', metric, 5)]
    opt_e = scenario_values[('Optimistic', metric, 5)]
    cons_e = scenario_values[('Conservative', metric, 5)]
    
    def fmt(v):
        if v is None: return 'None'
        if isinstance(v, (int, float)) and abs(v) < 2: return f"{v*100:.1f}%"
        return f"{v:,.0f}" if isinstance(v, (int, float)) else str(v)
    
    print(f"  {metric:<23s} {fmt(base_e):>12s} (r{rows['base']}) {fmt(opt_e):>12s} (r{rows['opt']}) {fmt(cons_e):>12s} (r{rows['cons']})")

# Now simulate: for each scenario, what SHOULD the Assumptions cells evaluate to?
print("\n" + "=" * 100)
print("SIMULATION: Expected Assumptions values when scenario is changed")
print("=" * 100)

for scenario_name in ['Base Case', 'Optimistic', 'Conservative']:
    print(f"\n  --- When Dashboard!B6 = '{scenario_name}' ---")
    for metric, rows in KEY_MAPPING.items():
        # The IF formula is: IF($B$64="Base Case", base, IF($B$64="Optimistic", opt, cons))
        expected_key = 'base' if scenario_name == 'Base Case' else ('opt' if scenario_name == 'Optimistic' else 'cons')
        expected_val = scenario_values[(scenario_name, metric, 5)]  # Col E
        
        if expected_val is None:
            print(f"    {metric:<25s}: None (EMPTY!)")
        elif isinstance(expected_val, (int, float)):
            if abs(expected_val) < 2:
                print(f"    {metric:<25s}: {expected_val*100:.1f}%")
            else:
                print(f"    {metric:<25s}: {expected_val:,.0f}")
        else:
            print(f"    {metric:<25s}: {expected_val}")

# Final check: verify the formula structure is correct
print("\n" + "=" * 100)
print("FORMULA STRUCTURE VERIFICATION")
print("=" * 100)

errors = []
for metric, rows in KEY_MAPPING.items():
    formula = asheet.cell(rows['assum'], 5).value  # E column
    if not formula or not isinstance(formula, str):
        errors.append(f"  {metric} E{rows['assum']}: NOT A FORMULA -> {repr(formula)}")
        continue
    
    # Check it references the correct base, opt, cons rows
    expected_base = f"Scenarios!E{rows['base']}"
    expected_opt = f"Scenarios!E{rows['opt']}"
    expected_cons = f"Scenarios!E{rows['cons']}"
    
    ok = True
    for exp in [expected_base, expected_opt, expected_cons]:
        if exp not in formula:
            errors.append(f"  {metric}: Missing reference {exp} in formula: {formula}")
            ok = False
    
    if ok:
        print(f"  {metric:<25s}: OK - refs {expected_base}, {expected_opt}, {expected_cons}")

if errors:
    print(f"\n  ERRORS ({len(errors)}):")
    for e in errors:
        print(f"    {e}")
else:
    print(f"\n  ALL 9 formulas reference correct rows - STRUCTURE IS CORRECT")

# Check if there are NON-key rows that are still hardcoded
print("\n" + "=" * 100)
print("NON-KEY ROWS CHECK: Are other Assumptions rows also formula-wired?")
print("=" * 100)
other_hardcoded = []
other_formula = []
for r in range(2, asheet.max_row):
    label = asheet.cell(r, 1).value
    if not label: continue
    
    for c in range(5, 10):
        v = asheet.cell(r, c).value
        if v and isinstance(v, str) and 'IF' in v:
            other_formula.append(r)
            break
        elif isinstance(v, (int, float)):
            if r not in [rows['assum'] for rows in KEY_MAPPING.values()]:
                other_hardcoded.append((r, str(label)[:40]))
            break

unique_hardcoded = list(set(other_hardcoded))
unique_hardcoded.sort()
print(f"  Rows with IF formulas: {len(set(other_formula))}")
print(f"  Rows with hardcoded numbers (non-key): {len(unique_hardcoded)}")
if unique_hardcoded:
    print("  Hardcoded rows:")
    for r, label in unique_hardcoded[:15]:
        v = asheet.cell(r, 5).value
        print(f"    Row {r}: {label} -> E{r}={v}")

wb.close()
wb_data.close()
print("\n=== Simulation Complete ===")
