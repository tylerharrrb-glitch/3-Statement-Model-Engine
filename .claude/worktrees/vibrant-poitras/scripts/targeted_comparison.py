"""
TARGETED COMPARISON: Read the Scenarios sheet's Optimistic block
and compare against what the financial statement formulas would compute.

The key question: do the IS/BS/CF formulas INDEPENDENTLY produce the same
values as the engine-computed values stored in the Scenarios sheet?

If they do → the fix just needs correct scenario data (my latest fix handles this)
If they don't → the formulas are wrong and need to be fixed
"""
import openpyxl
from openpyxl.utils import get_column_letter
import os, sys, re

sys.stdout.reconfigure(encoding='utf-8')

SRC = os.path.join(os.environ['USERPROFILE'], 'Downloads', 'WOLF_Financial_Model.xlsx')

wb = openpyxl.load_workbook(SRC)
wb_d = openpyxl.load_workbook(SRC, data_only=True)

print("FILE:", SRC)
print("Modified:", os.path.getmtime(SRC))

# ═══════════════════════════════════════════════
# A) Check the Scenarios sheet — do all 3 blocks have DIFFERENT values?
# ═══════════════════════════════════════════════
scen = wb_d['Scenarios']

# Find block starts
block_starts = {}
for r in range(1, scen.max_row + 1):
    val = str(scen.cell(r, 1).value or '').upper()
    if 'BASE CASE' in val and 'OPTIMISTIC' not in val and 'CONSERVATIVE' not in val:
        block_starts['BASE'] = r
    elif 'OPTIMISTIC' in val:
        block_starts['OPT'] = r
    elif 'CONSERVATIVE' in val:
        block_starts['CONS'] = r

print("\n═══════════════════════════════════════════════")
print("A) SCENARIO BLOCKS — Row mapping")
print("═══════════════════════════════════════════════")
for name, start in block_starts.items():
    print(f"  {name}: starts at row {start}")

# For each block, extract the key data rows (relative offsets from block start)
def extract_block(start_row, num_rows=62):
    """Extract all data from a block."""
    data = {}
    for r in range(start_row, start_row + num_rows):
        label = scen.cell(r, 1).value
        if not label: continue
        label = str(label).strip()
        vals = []
        for c in range(2, 10):  # cols B through I
            v = scen.cell(r, c).value
            vals.append(v)
        data[label] = {'row': r, 'vals': vals}
    return data

base_data = extract_block(block_starts['BASE'])
opt_data = extract_block(block_starts['OPT'])
cons_data = extract_block(block_starts['CONS'])

# Now compare key metrics across scenarios
print("\n═══════════════════════════════════════════════")
print("B) KEY METRIC COMPARISON (Projection year 5 = 2028E, ~col I)")
print("═══════════════════════════════════════════════")

key_metrics = [
    'Revenue Growth Rate', 'COGS % of Revenue', 'SG&A % of Revenue',
    'R&D % of Revenue', 'CapEx % of Revenue', 'Interest Rate (on Debt)',
    'Dividend Payout Ratio', 'Other OpEx % of Revenue', 'Tax Rate',
    'Interest Income (Computed)', 'Interest Expense (Computed)',
    'Depreciation (Computed)', 'Net PP&E (Computed)',
    'Retained Earnings (Computed)', 'Dividends Paid (Computed)',
    'Long-Term Debt (Computed)',
]

for metric in key_metrics:
    b = base_data.get(metric, {}).get('vals', [])
    o = opt_data.get(metric, {}).get('vals', [])
    c = cons_data.get(metric, {}).get('vals', [])
    
    # Get the last projection value (index 6 = col I = 2028E)
    def fmt(v):
        if v is None: return 'N/A'
        if isinstance(v, (int, float)):
            if abs(v) < 2 and abs(v) > 0:
                return f"{v*100:.2f}%"
            return f"{v:>12,.0f}"
        return str(v)[:12]
    
    # Show all projection values (indices 3-7 = cols E-I = 2024E-2028E)
    b_val = fmt(b[6]) if len(b) > 6 else 'N/A'
    o_val = fmt(o[6]) if len(o) > 6 else 'N/A'
    c_val = fmt(c[6]) if len(c) > 6 else 'N/A'
    
    same = ' ⚠️ SAME!' if b_val == o_val and b_val == c_val else ''
    print(f"  {metric:40s}  Base={b_val:>12s}  Opt={o_val:>12s}  Cons={c_val:>12s}{same}")

# ═══════════════════════════════════════════════
# C) Full Optimistic block dump vs Base Case to spot differences
# ═══════════════════════════════════════════════
print("\n═══════════════════════════════════════════════")
print("C) DIFFERENCES: Base vs Optimistic (2024E through 2028E)")
print("═══════════════════════════════════════════════")

for label in base_data:
    if label in opt_data:
        b_vals = base_data[label]['vals']
        o_vals = opt_data[label]['vals']
        
        # Compare projection columns (indices 3-7)
        diffs = []
        for i in range(3, min(len(b_vals), len(o_vals), 8)):
            bv = b_vals[i] if i < len(b_vals) else None
            ov = o_vals[i] if i < len(o_vals) else None
            if bv != ov and bv is not None and ov is not None:
                diffs.append(f"col{i+2}: {bv:>10,.0f} → {ov:>10,.0f}" if isinstance(bv, (int, float)) else f"col{i+2}: {bv} → {ov}")
        
        if diffs:
            print(f"  {label:40s}  {'; '.join(diffs[:3])}")

# ═══════════════════════════════════════════════
# D) Check IF formula structure — count unique column references
# ═══════════════════════════════════════════════
print("\n═══════════════════════════════════════════════")
print("D) BALANCE SHEET FORMULAS — Architecture Check")
print("═══════════════════════════════════════════════")

bs_ws = wb['Balance Sheet']
bs_ws_d = wb_d['Balance Sheet']

# Show all BS formulas for projection column E
for r in range(1, bs_ws.max_row + 1):
    label = bs_ws.cell(r, 1).value
    if not label: continue
    label = str(label).strip()
    
    # Check col E formula
    formula = bs_ws.cell(r, 5).value  # Column E = first projection
    val = bs_ws_d.cell(r, 5).value
    
    if formula and isinstance(formula, str):
        val_str = f"{val:>12,.0f}" if isinstance(val, (int, float)) else str(val)
        print(f"  Row {r:3d}: {label:35s}  val={val_str}  formula={str(formula)[:80]}")

# ═══════════════════════════════════════════════
# E) Check Cash Flow formulas  
# ═══════════════════════════════════════════════
print("\n═══════════════════════════════════════════════")
print("E) CASH FLOW FORMULAS — Full Architecture (Col E)")
print("═══════════════════════════════════════════════")

cf_ws = wb['Cash Flow Statement']
cf_ws_d = wb_d['Cash Flow Statement']

for r in range(1, cf_ws.max_row + 1):
    label = cf_ws.cell(r, 1).value
    if not label: continue
    label = str(label).strip()
    
    formula = cf_ws.cell(r, 5).value  # Column E
    val = cf_ws_d.cell(r, 5).value
    
    if formula:
        val_str = f"{val:>12,.0f}" if isinstance(val, (int, float)) else str(val)
        f_str = str(formula)[:90]
        print(f"  Row {r:3d}: {label:35s}  val={val_str}  formula={f_str}")

wb.close()
wb_d.close()
print("\n=== Analysis Complete ===")
