"""
Full diagnostic on the fresh WOLF export — checks all scenario wiring.
"""
import openpyxl
from openpyxl.utils import get_column_letter
import os, sys

sys.stdout.reconfigure(encoding='utf-8')

SRC = os.path.join(os.environ['USERPROFILE'], 'Downloads', 'WOLF_Financial_Model.xlsx')
print(f"Checking: {SRC}")
print(f"File size: {os.path.getsize(SRC):,} bytes")
print(f"Modified: {os.path.getmtime(SRC)}")

# ── Sheet Order ──
wb = openpyxl.load_workbook(SRC)
print(f"\nSheet order: {wb.sheetnames}")
dash_idx = wb.sheetnames.index('Dashboard') if 'Dashboard' in wb.sheetnames else -1
print(f"Dashboard position: {dash_idx} (should be 0)")

# ── Scenarios Sheet ──
print("\n" + "=" * 70)
print("SCENARIOS SHEET")
print("=" * 70)
scen = wb['Scenarios']
blocks = []
for r in range(1, scen.max_row + 1):
    v = str(scen.cell(r, 1).value or '')
    if '\u258e' in v:
        blocks.append((r, v.replace('\u258e', '|')))
        print(f"  Block at row {r}: {v.replace(chr(0x258e), '|')}")

print(f"  Total blocks: {len(blocks)}")

# ── Dashboard B6 ──
print("\n" + "=" * 70)
print("DASHBOARD B6")
print("=" * 70)
dash = wb['Dashboard']
print(f"  Value: {repr(dash['B6'].value)}")
for dv in dash.data_validations.dataValidation:
    for cr in dv.cells.ranges:
        if 'B6' in str(cr):
            print(f"  Validation: type={dv.type}, formula1={dv.formula1}")

# ── Control Cell ──
print("\n" + "=" * 70)
print("ASSUMPTIONS CONTROL CELL")
print("=" * 70)
asheet = wb['Assumptions']
control_addr = None
for r in range(1, asheet.max_row + 1):
    for c in range(1, 5):
        v = asheet.cell(r, c).value
        if v and 'Dashboard' in str(v):
            addr = f"{get_column_letter(c)}{r}"
            print(f"  Control at {addr}: {repr(v)}")
            control_addr = addr

# ── IF Formulas ──
print("\n" + "=" * 70)
print("IF FORMULAS IN ASSUMPTIONS (projection cols E-I)")
print("=" * 70)

TARGET_LABELS = [
    'Revenue Growth Rate', 'COGS % of Revenue', 'SG&A % of Revenue',
    'R&D % of Revenue', 'Other OpEx % of Revenue', 'Tax Rate',
    'CapEx % of Revenue', 'Interest Rate', 'Dividend Payout',
]

found_rows = {}
if_formula_count = 0
hardcoded_count = 0

for r in range(1, asheet.max_row + 1):
    label = str(asheet.cell(r, 1).value or '')
    for t in TARGET_LABELS:
        if t.lower() in label.lower() and t not in found_rows:
            found_rows[t] = r

for t in TARGET_LABELS:
    r = found_rows.get(t)
    if not r:
        print(f"  NOT FOUND: {t}")
        continue
    label = asheet.cell(r, 1).value
    has_if = False
    for c in range(5, 10):
        v = asheet.cell(r, c).value
        if v and isinstance(v, str) and 'IF' in v.upper():
            if_formula_count += 1
            has_if = True
        elif isinstance(v, (int, float)):
            hardcoded_count += 1
    
    if has_if:
        sample = asheet.cell(r, 5).value
        print(f"  Row {r} {str(label)[:35]:35s}: IF formulas (5 cols)")
        print(f"    Sample E{r}: {str(sample)[:100]}")
    else:
        vals = [f"{asheet.cell(r, c).value}" for c in range(5, 10)]
        print(f"  Row {r} {str(label)[:35]:35s}: HARDCODED → {', '.join(vals)}")

# ── Scenario Data Comparison ──
print("\n" + "=" * 70)
print("SCENARIO DATA COMPARISON")
print("=" * 70)

# Read data-only version
wb_data = openpyxl.load_workbook(SRC, data_only=True)
scen_data = wb_data['Scenarios']

# Find data rows for each block by reading IF formulas
# Extract row numbers from the E5 formula
e5_formula = asheet.cell(found_rows.get('Revenue Growth Rate', 5), 5).value
if e5_formula and isinstance(e5_formula, str):
    print(f"\n  E5 formula: {e5_formula}")
    # Parse out the Scenarios! references
    import re
    refs = re.findall(r'Scenarios!([A-Z])(\d+)', e5_formula)
    if refs:
        print(f"  References: {refs}")
        for col, row_num in refs:
            r = int(row_num)
            vals = []
            for c in range(5, 10):
                v = scen_data.cell(r, c).value
                if isinstance(v, (int, float)) and abs(v) < 2:
                    vals.append(f"{v*100:.1f}%")
                elif isinstance(v, (int, float)):
                    vals.append(f"{v:,.0f}")
                else:
                    vals.append(str(v))
            print(f"  Scenarios row {r}: {', '.join(vals)}")

# Total IF formula count (all rows, not just 9 key ones)
total_if = 0
total_hardcoded_proj = 0
for r in range(2, asheet.max_row + 1):
    for c in range(5, 10):
        v = asheet.cell(r, c).value
        if v and isinstance(v, str) and 'IF' in v:
            total_if += 1
        elif isinstance(v, (int, float)):
            total_hardcoded_proj += 1

print(f"\n  Total IF formulas in ALL Assumptions rows: {total_if}")
print(f"  Total hardcoded values in projection cols: {total_hardcoded_proj}")

# ── Summary ──
print("\n" + "=" * 70)
print("FINAL SUMMARY")
print("=" * 70)
print(f"  Dashboard first: {'YES' if dash_idx == 0 else 'NO'}")
print(f"  Scenario blocks: {len(blocks)}/3")
print(f"  IF formulas (9 key rows): {if_formula_count}/45")
print(f"  Hardcoded (9 key rows): {hardcoded_count}")
print(f"  Control cell: {control_addr}")

all_good = len(blocks) == 3 and if_formula_count == 45 and dash_idx == 0 and control_addr
if all_good:
    print("\n  *** ALL CHECKS PASSED ***")
else:
    issues = []
    if len(blocks) < 3: issues.append(f"Only {len(blocks)} blocks")
    if if_formula_count < 45: issues.append(f"Only {if_formula_count}/45 IF formulas")
    if dash_idx != 0: issues.append("Dashboard not first")
    if not control_addr: issues.append("No control cell")
    print(f"\n  *** ISSUES: {'; '.join(issues)} ***")

wb.close()
wb_data.close()
