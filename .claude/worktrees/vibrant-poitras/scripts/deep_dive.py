"""
DEEP DIVE: Check if the Scenarios sheet has DIFFERENT computed values
for Base vs Optimistic vs Conservative.
The hypothesis is that some computed rows are identical across scenarios,
meaning the engine didn't properly recompute them.
"""
import openpyxl
import os, sys

sys.stdout.reconfigure(encoding='utf-8')

SRC = os.path.join(os.environ['USERPROFILE'], 'Downloads', 'WOLF_Financial_Model.xlsx')
wb_d = openpyxl.load_workbook(SRC, data_only=True)
scen = wb_d['Scenarios']

# Find block starts
blocks = {}
for r in range(1, scen.max_row + 1):
    val = str(scen.cell(r, 1).value or '').upper()
    if 'BASE CASE' in val and 'OPTIMISTIC' not in val and 'CONSERVATIVE' not in val:
        blocks['BASE'] = r
    elif 'OPTIMISTIC' in val:
        blocks['OPT'] = r
    elif 'CONSERVATIVE' in val:
        blocks['CONS'] = r

print("Block starts:", blocks)
print()

# For each block, get all labeled rows and their values
def get_block_rows(start, count=62):
    rows = {}
    for r in range(start, start + count):
        label = scen.cell(r, 1).value
        if not label: continue
        label = str(label).strip()
        vals = []
        for c in range(2, 10):
            v = scen.cell(r, c).value
            vals.append(v)
        rows[label] = vals
    return rows

base = get_block_rows(blocks['BASE'])
opt = get_block_rows(blocks['OPT'])
cons = get_block_rows(blocks['CONS'])

# Compare all labels
print("=" * 120)
print("EVERY ROW: Base vs Optimistic vs Conservative (projection cols only)")
print("=" * 120)

all_labels = list(base.keys())
for label in all_labels:
    if label.startswith('──') or label in ['Year', 'Period']:
        continue
    b = base.get(label, [])
    o = opt.get(label, [])
    c = cons.get(label, [])
    
    # Projection cols are typically indices 3-7 (E through I)
    def get_proj(vals, idx_start=3, idx_end=8):
        return [vals[i] if i < len(vals) else None for i in range(idx_start, idx_end)]
    
    b_proj = get_proj(b)
    o_proj = get_proj(o)
    c_proj = get_proj(c)
    
    # Check if all same
    all_same = (b_proj == o_proj == c_proj)
    
    def fmt_val(v):
        if v is None: return 'N/A'
        if isinstance(v, (int, float)):
            if abs(v) < 2 and v != 0:
                return f"{v*100:.2f}%"
            return f"{v:>12,.0f}"
        return str(v)[:12]
    
    # Only show projection year 1 (col E = index 3) as representative
    b_str = fmt_val(b_proj[0])
    o_str = fmt_val(o_proj[0])
    c_str = fmt_val(c_proj[0])
    
    marker = '  SAME' if all_same else '  DIFF ✓'
    if all_same and b_proj[0] is not None and isinstance(b_proj[0], (int, float)) and abs(b_proj[0]) > 0.001:
        marker = '  ⚠️ SAME!'  # These SHOULD differ but don't
    
    print(f"  {label:45s}  Base={b_str:>12s}  Opt={o_str:>12s}  Cons={c_str:>12s}{marker}")

# Also check the Assumptions sheet's IF formula references
print("\n" + "=" * 120)
print("ASSUMPTIONS SHEET: Which rows have IF formulas AND which don't?")
print("=" * 120)

asheet = wb_d['Assumptions']
asheet_f = wb_d['Assumptions']  # Need formula version
asheet_f2 = openpyxl.load_workbook(SRC)['Assumptions']

for r in range(1, asheet.max_row + 1):
    label = asheet.cell(r, 1).value
    if not label: continue
    label = str(label).strip()
    
    # Check projection cols (E=5 through I=9) for IF formulas
    has_if = False
    has_hardcoded = False
    sample_formula = None
    sample_val = None
    
    for c in range(5, 10):
        v = asheet_f2.cell(r, c).value
        if v and isinstance(v, str) and '=' in v and 'IF' in v:
            has_if = True
            if not sample_formula:
                sample_formula = str(v)[:60]
        elif isinstance(v, (int, float)):
            has_hardcoded = True
            if sample_val is None:
                sample_val = v
    
    if has_if and has_hardcoded:
        print(f"  Row {r:3d}: {label:40s}  MIXED ⚠️  IF+hardcoded  sample_val={sample_val}")
    elif has_if:
        pass  # Expected - has IF formulas  
    elif has_hardcoded:
        val_str = fmt_val(sample_val)
        print(f"  Row {r:3d}: {label:40s}  HARDCODED ⚠️  val={val_str}")
    elif label and any(c in label.lower() for c in ['(', 'section', '──']):
        pass  # Section header
    elif label.strip():
        # Check if it has any value at all
        any_val = any(asheet.cell(r, c).value is not None for c in range(5, 10))
        if any_val:
            print(f"  Row {r:3d}: {label:40s}  NO FORMULA  (has values)")

wb_d.close()
print("\n=== Deep Dive Complete ===")
