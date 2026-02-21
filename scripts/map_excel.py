"""
Step 1-5: Map the existing Excel workbook structure.
This reads the WOLF file (clean export) since the v3 may already be modified.
"""
import openpyxl
import sys
import os

# Use WOLF file as the clean base (same structure as Demo)
SRC = r"C:\Users\user\Downloads\WOLF_Financial_Model.xlsx"
if not os.path.exists(SRC):
    SRC = r"C:\Users\user\Downloads\Demo_Company_Inc__Financial_Model_v3.xlsx"

print(f"=== MAPPING: {os.path.basename(SRC)} ({os.path.getsize(SRC)} bytes) ===\n")

# Step 1: data_only=True - get computed values
print("=" * 80)
print("STEP 1: SHEET OVERVIEW (data_only=True)")
print("=" * 80)
wb_val = openpyxl.load_workbook(SRC, data_only=True)
for name in wb_val.sheetnames:
    ws = wb_val[name]
    print(f"\n--- Sheet: '{name}' | Dims: {ws.dimensions} ---")
    for r in range(1, min(6, ws.max_row + 1)):
        cols = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None:
                cols.append(f"{openpyxl.utils.get_column_letter(c)}{r}={repr(v)[:60]}")
        if cols:
            print(f"  Row {r}: {', '.join(cols)}")
wb_val.close()

# Step 2: data_only=False - get formulas
print("\n" + "=" * 80)
print("STEP 2: FULL FORMULA DUMP FOR KEY SHEETS")
print("=" * 80)
wb = openpyxl.load_workbook(SRC, data_only=False)

for name in ['Assumptions', 'Income Statement', 'Balance Sheet', 'Cash Flow Statement', 'Ratios']:
    if name not in wb.sheetnames:
        print(f"\n*** Sheet '{name}' NOT FOUND ***")
        continue
    ws = wb[name]
    print(f"\n{'='*60}")
    print(f"SHEET: '{name}' | rows={ws.max_row}, cols={ws.max_column}")
    print(f"{'='*60}")
    for r in range(1, ws.max_row + 1):
        cells = []
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            v = cell.value
            if v is not None:
                col_l = openpyxl.utils.get_column_letter(c)
                cells.append(f"{col_l}={repr(v)[:80]}")
        if cells:
            print(f"  R{r:02d}: {' | '.join(cells)}")

# Also dump any other sheets
for name in wb.sheetnames:
    if name in ['Assumptions', 'Income Statement', 'Balance Sheet', 'Cash Flow Statement', 'Ratios']:
        continue
    ws = wb[name]
    print(f"\n{'='*60}")
    print(f"SHEET: '{name}' | rows={ws.max_row}, cols={ws.max_column}")
    print(f"{'='*60}")
    for r in range(1, min(ws.max_row + 1, 60)):
        cells = []
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            v = cell.value
            if v is not None:
                col_l = openpyxl.utils.get_column_letter(c)
                cells.append(f"{col_l}={repr(v)[:80]}")
        if cells:
            print(f"  R{r:02d}: {' | '.join(cells)}")

wb.close()

# Step 3: Record key assumption row numbers  
print("\n" + "=" * 80)
print("STEP 3: ASSUMPTION ROW MAP")
print("=" * 80)
wb = openpyxl.load_workbook(SRC, data_only=False)
ws = wb['Assumptions']
assumption_keys = [
    'Revenue Growth', 'COGS', 'SG&A', 'R&D', 'Other OpEx', 'Tax Rate',
    'CapEx', 'Interest Rate', 'Dividend Payout'
]
for r in range(1, ws.max_row + 1):
    label = str(ws.cell(r, 1).value or '')
    for key in assumption_keys:
        if key.lower() in label.lower():
            vals = [ws.cell(r, c).value for c in range(2, 10)]
            print(f"  Row {r}: '{label}' -> {vals}")
            break
wb.close()

print("\n=== MAPPING COMPLETE ===")
