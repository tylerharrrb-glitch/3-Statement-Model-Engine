"""Build the Dashboard sheet."""
from enhance_helpers import *
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation

def build_dashboard(wb):
    ws = wb.create_sheet('Dashboard', 0)
    ws.sheet_properties.tabColor = NAVY
    for c in range(1, 11):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.column_dimensions['A'].width = 24
    AL = Alignment(vertical='center', wrap_text=True)
    ALC = Alignment(vertical='center', horizontal='center')
    ALR = Alignment(vertical='center', horizontal='right')

    # ── Section A: Company Banner ─────────────────
    merge_and_style(ws, 1, 1, 10, "='Company Info'!B4",
                    ft=font(WHITE, 22, True), fl=fill(NAVY), align=ALC)
    ws.row_dimensions[1].height = 48
    merge_and_style(ws, 2, 1, 10,
        'FINANCIAL MODEL DASHBOARD  ·  E£ (Egyptian Pound)  ·  February 20, 2026  ·  CONFIDENTIAL',
        ft=font(WHITE, 11), fl=fill(MED_BLUE), align=ALC)
    ws.row_dimensions[2].height = 22
    merge_and_style(ws, 3, 1, 10, '', fl=fill(LIGHT_BLUE))
    ws.row_dimensions[3].height = 6

    # ── Section B: Scenario Control ───────────────
    merge_and_style(ws, 4, 1, 10, '\U0001F4CA  SCENARIO CONTROL',
                    ft=font(WHITE, 12, True), fl=fill(MED_BLUE), align=AL)

    # Row 5: labels
    style_cell(ws, 5, 1, 'Active Scenario', ft=font(NAVY, 10, True), fl=fill(LIGHT_GRAY), align=AL, border=thin_border)
    merge_and_style(ws, 5, 2, 4, 'Select Scenario →', ft=font(AMBER, 10, italic=True), fl=fill(LIGHT_AMBER), align=ALC)
    style_cell(ws, 5, 5, 'Revenue Base (E£)', ft=font(NAVY, 10, True), fl=fill(LIGHT_GRAY), align=AL, border=thin_border)
    c = style_cell(ws, 5, 6, "='Company Info'!B18", ft=font(DARK_GRAY, 10), fl=fill(WHITE), align=ALR, border=thin_border)
    c.number_format = '#,##0'
    style_cell(ws, 5, 7, 'Fiscal Year', ft=font(NAVY, 10, True), fl=fill(LIGHT_GRAY), align=AL, border=thin_border)
    style_cell(ws, 5, 8, "='Company Info'!B15", ft=font(DARK_GRAY, 10), fl=fill(WHITE), align=ALC, border=thin_border)

    # Row 6: Dropdown
    style_cell(ws, 6, 1, '▼ Scenario Selector', ft=font(WHITE, 11, True), fl=fill(NAVY), align=AL, border=thin_border)
    merge_and_style(ws, 6, 2, 4, 'Base Case', ft=font(NAVY, 13, True), fl=fill(LIGHT_BLUE), align=ALC)
    # Apply thick borders to merged dropdown cell
    for c in range(2, 5):
        ws.cell(6, c).border = thick_blue_border
    dv = DataValidation(type='list', formula1='"Base Case,Optimistic,Conservative"', showDropDown=False)
    dv.promptTitle = 'Scenario Selection'
    dv.prompt = 'Choose Base Case, Optimistic, or Conservative scenario'
    dv.errorTitle = 'Invalid Scenario'
    dv.error = 'Please select a valid scenario from the list'
    ws.add_data_validation(dv)
    dv.add('B6')

    # Description formula
    desc_formula = ('=IF(B6="Base Case","Central planning — stable market conditions  |  '
                   'Revenue growth 10%→5%  |  COGS 60%  |  Rate 20%",'
                   'IF(B6="Optimistic","Bull case — strong execution  |  '
                   'Revenue growth 15%→8%  |  COGS 55%  |  Rate 15%",'
                   '"Bear case — headwinds  |  Revenue growth 5%→2%  |  COGS 63%  |  Rate 22%"))')
    merge_and_style(ws, 6, 5, 10, desc_formula,
                    ft=font(DARK_GRAY, 9, italic=True), fl=fill(WHITE), align=AL)

    # Row 7: spacer
    merge_and_style(ws, 7, 1, 10, '', fl=fill(LIGHT_BLUE))
    ws.row_dimensions[7].height = 6

    # ── Section C: Key Metrics ────────────────────
    merge_and_style(ws, 8, 1, 10, '\U0001F4C8  KEY METRICS — ACTIVE SCENARIO',
                    ft=font(WHITE, 12, True), fl=fill(NAVY), align=AL)

    # Row 9: Period headers
    headers = ['', '2021 (A)', '2022 (A)', '2023 (A)', '2024E', '2025E', '2026E', '2027E', '2028E', '']
    for ci, h in enumerate(headers):
        style_cell(ws, 9, ci+1, h, ft=font(WHITE, 10, True), fl=fill(NAVY), align=ALC, border=thin_border)

    # ── Metric rows ──────────────────────────────
    # IS_map: dash_col → IS_col. B→B(2021), C→C(2022), D→D(2023), E→E(2024E)...I→I(2028E)
    # CF_map: CF starts at 2022. B→none, C→B(2022), D→C(2023), E→D(2024E)...I→H(2028E)
    # BS: same as IS

    def is_ref(col, row):
        """IS ref: Dashboard col → IS col (same letters)"""
        return f"='Income Statement'!{col}{row}"
    def bs_ref(col, row):
        return f"='Balance Sheet'!{col}{row}"
    def cf_ref(dash_col, cf_row):
        """CF ref: Dashboard col B→none, C→B, D→C, E→D, F→E, G→F, H→G, I→H"""
        cf_map = {'B': None, 'C': 'B', 'D': 'C', 'E': 'D', 'F': 'E', 'G': 'F', 'H': 'G', 'I': 'H'}
        cf_col = cf_map.get(dash_col)
        if cf_col is None:
            return '—'  # No 2021 CF data
        return f"='Cash Flow Statement'!{cf_col}{cf_row}"
    def ratio_ref(col, row):
        return f"=Ratios!{col}{row}"

    cols = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
    
    # Define metric rows: (label, formulas_func, number_format, bold, section)
    metrics = [
        # Income Statement section
        ('section', '\U0001F4D7  INCOME STATEMENT', LIGHT_BLUE),
        ('Revenue (E£)', lambda c: is_ref(c, 2), '#,##0', True),
        ('Revenue Growth %', lambda c: is_ref(c, 3), '0.0%', False),
        ('Gross Profit (E£)', lambda c: is_ref(c, 5), '#,##0', False),
        ('Gross Margin %', lambda c: is_ref(c, 6), '0.0%', False),
        ('EBITDA (E£)', lambda c: is_ref(c, 18), '#,##0', False),
        ('EBITDA Margin %', lambda c: f"=IF('Income Statement'!{c}2=0,0,'Income Statement'!{c}18/'Income Statement'!{c}2)", '0.0%', False),
        ('EBIT (E£)', lambda c: is_ref(c, 17), '#,##0', False),
        ('EBIT Margin %', lambda c: is_ref(c, 19), '0.0%', False),
        ('Net Income (E£)', lambda c: is_ref(c, 28), '#,##0', True),
        ('Net Margin %', lambda c: is_ref(c, 29), '0.0%', False),
        ('EPS (E£)', lambda c: is_ref(c, 30), '#,##0.00', False),
        # Cash Flow section
        ('section', '\U0001F4B5  CASH FLOW', LIGHT_BLUE),
        ('Cash from Operations (E£)', lambda c: cf_ref(c, 16), '#,##0', False),
        ('Free Cash Flow (E£)', lambda c: cf_ref(c, 36), '#,##0', True),
        ('Ending Cash (E£)', lambda c: cf_ref(c, 34), '#,##0', False),
        # Balance Sheet section
        ('section', '\U0001F3E6  BALANCE SHEET', LIGHT_BLUE),
        ('Total Assets (E£)', lambda c: bs_ref(c, 15), '#,##0', True),
        ('Total Debt (E£)', lambda c: f"='Balance Sheet'!{c}20+'Balance Sheet'!{c}21+'Balance Sheet'!{c}26", '#,##0', False),
        ('Total Equity (E£)', lambda c: bs_ref(c, 38), '#,##0', True),
        ('Cash & Equivalents (E£)', lambda c: bs_ref(c, 3), '#,##0', False),
        # Ratios section
        ('section', '\U0001F4CA  KEY RATIOS', LIGHT_BLUE),
        ('Current Ratio', lambda c: ratio_ref(c, 9), '0.00"x"', False),
        ('Debt / Equity', lambda c: ratio_ref(c, 10), '0.00"x"', False),
        ('ROE %', lambda c: ratio_ref(c, 7), '0.0%', False),
        ('ROA %', lambda c: ratio_ref(c, 6), '0.0%', False),
    ]

    row = 10
    alt = 0
    for m in metrics:
        if m[0] == 'section':
            merge_and_style(ws, row, 1, 10, m[1],
                           ft=font(NAVY, 10, True), fl=fill(m[2]), align=AL)
            row += 1
            alt = 0
            continue
        label, formula_fn, nfmt, bold_flag = m[0], m[1], m[2], m[3]
        bg = LIGHT_GRAY if alt % 2 == 0 else WHITE
        ws.row_dimensions[row].height = 18
        style_cell(ws, row, 1, label, ft=font(NAVY, 10, bold_flag), fl=fill(bg), align=AL, border=thin_border)
        for ci, col_letter in enumerate(cols):
            val = formula_fn(col_letter)
            c = style_cell(ws, row, 2+ci, val,
                          ft=font(DARK_GRAY if not bold_flag else NAVY, 10, bold_flag),
                          fl=fill(bg), align=ALR, border=thin_border)
            if val != '—':
                c.number_format = nfmt
        style_cell(ws, row, 10, '', fl=fill(bg), border=thin_border)
        row += 1
        alt += 1

    # ── Spacer ────────────────────────────────────
    row += 1
    merge_and_style(ws, row, 1, 10, '', fl=fill(LIGHT_BLUE))
    ws.row_dimensions[row].height = 6
    row += 1

    # ── Section D: Scenario Comparison ────────────
    merge_and_style(ws, row, 1, 10, '⚖️  SCENARIO COMPARISON — Terminal Year 2028E',
                    ft=font(WHITE, 12, True), fl=fill(NAVY), align=AL)
    row += 1
    merge_and_style(ws, row, 1, 10,
        'Values are pre-computed scenario snapshots. Change the active scenario via the dropdown above.',
        ft=font(AMBER, 9, italic=True), fl=fill(LIGHT_AMBER), align=AL)
    row += 1

    # Column headers
    comp_headers = ['Metric', '', 'Base Case', '', 'Optimistic', '', 'Conservative', '', 'Base→Opt Δ', 'Base→Cons Δ']
    for ci, h in enumerate(comp_headers):
        bg_h = NAVY if h else MED_GRAY
        style_cell(ws, row, ci+1, h, ft=font(WHITE, 10, True), fl=fill(bg_h), align=ALC, border=thin_border)
    row += 1

    # Pre-computed 2028E values from mapping data (WOLF model base case values)
    # These are static snapshots calculated from each scenario
    from build_scenarios import BASE_DATA, OPT_DATA, CONS_DATA

    # We need to compute approximate 2028E values for each scenario
    # Revenue chain: base=1000000, then 5 years of growth
    def compute_rev(data):
        r = 1000000
        for g in data[0]:
            r *= (1 + g)
        return r

    def compute_metrics(data):
        rev = compute_rev(data)
        cogs = rev * data[1][4]
        gp = rev - cogs
        sga = rev * data[2][4]
        rd = rev * data[3][4]
        dep = 65000  # approximate
        amort = 5000
        other = rev * data[4][4]
        sbc = 10000
        tot_opex = sga + rd + dep + amort + other + sbc
        ebit = gp - tot_opex
        ebitda = ebit + dep + amort
        int_exp = 180000 * data[7][4]  # approx total debt * rate
        int_inc = 3000
        ebt = ebit + int_inc - int_exp
        tax = max(0, ebt * data[5][4])
        ni = ebt - tax
        capex = rev * data[6][4]
        cfo = ni + dep + amort + sbc
        fcf = cfo - capex
        gm = gp / rev if rev else 0
        ebitda_m = ebitda / rev if rev else 0
        nm = ni / rev if rev else 0
        return {'Revenue': rev, 'Gross Margin %': gm, 'EBITDA': ebitda,
                'EBITDA Margin %': ebitda_m, 'EBIT': ebit, 'Net Income': ni,
                'Net Margin %': nm, 'Free Cash Flow': fcf}

    base_m = compute_metrics(BASE_DATA)
    opt_m = compute_metrics(OPT_DATA)
    cons_m = compute_metrics(CONS_DATA)

    comp_metrics = ['Revenue', 'Gross Margin %', 'EBITDA', 'EBITDA Margin %',
                    'EBIT', 'Net Income', 'Net Margin %', 'Free Cash Flow']

    for i, metric in enumerate(comp_metrics):
        bg = LIGHT_GRAY if i % 2 == 0 else WHITE
        bv = base_m[metric]
        ov = opt_m[metric]
        cv = cons_m[metric]
        is_pct = '%' in metric
        nfmt = '0.0%' if is_pct else '#,##0'

        style_cell(ws, row, 1, f'{metric} (E£)' if not is_pct else metric,
                  ft=font(NAVY, 10, True), fl=fill(bg), align=AL, border=thin_border)
        style_cell(ws, row, 2, '', fl=fill(MED_GRAY), border=thin_border)
        c = style_cell(ws, row, 3, bv, ft=font(NAVY, 10, True), fl=fill(VERY_LIGHT_BLUE), align=ALR, border=thin_border)
        c.number_format = nfmt
        style_cell(ws, row, 4, '', fl=fill(MED_GRAY), border=thin_border)
        c = style_cell(ws, row, 5, ov, ft=font(DARK_GREEN, 10, True), fl=fill(LIGHT_GREEN), align=ALR, border=thin_border)
        c.number_format = nfmt
        style_cell(ws, row, 6, '', fl=fill(MED_GRAY), border=thin_border)
        c = style_cell(ws, row, 7, cv, ft=font(AMBER, 10, True), fl=fill(LIGHT_AMBER), align=ALR, border=thin_border)
        c.number_format = nfmt
        style_cell(ws, row, 8, '', fl=fill(MED_GRAY), border=thin_border)

        # Deltas
        if bv != 0:
            delta_opt = (ov - bv) / abs(bv) if not is_pct else ov - bv
            delta_cons = (cv - bv) / abs(bv) if not is_pct else cv - bv
        else:
            delta_opt = delta_cons = 0
        dfmt = '+0.0%;-0.0%'
        dc = style_cell(ws, row, 9, delta_opt, ft=font(GREEN if delta_opt >= 0 else RED, 10, True),
                        fl=fill(bg), align=ALC, border=thin_border)
        dc.number_format = dfmt
        dc = style_cell(ws, row, 10, delta_cons, ft=font(GREEN if delta_cons >= 0 else RED, 10, True),
                        fl=fill(bg), align=ALC, border=thin_border)
        dc.number_format = dfmt
        row += 1

    # Assumption summary sub-table
    row += 1
    merge_and_style(ws, row, 1, 10, '2028E Assumption Summary',
                    ft=font(WHITE, 10, True), fl=fill('555555'), align=ALC)
    row += 1
    summ_metrics = [
        ('Revenue Growth Rate', '5.0%', '8.0%', '2.0%'),
        ('COGS %', '60.0%', '55.0%', '62.0%'),
        ("SG&A %", '15.0%', '13.0%', '16.0%'),
        ('Interest Rate', '20.0%', '15.0%', '20.0%'),
        ('Dividend Payout', '5.0%', '12.0%', '3.0%'),
    ]
    for i, (label, bv, ov, cv) in enumerate(summ_metrics):
        bg = LIGHT_GRAY if i % 2 == 0 else WHITE
        style_cell(ws, row, 1, label, ft=font(NAVY, 9), fl=fill(bg), align=AL, border=thin_border)
        style_cell(ws, row, 2, '', fl=fill(MED_GRAY), border=thin_border)
        style_cell(ws, row, 3, bv, ft=font(NAVY, 9), fl=fill(VERY_LIGHT_BLUE), align=ALC, border=thin_border)
        style_cell(ws, row, 4, '', fl=fill(MED_GRAY), border=thin_border)
        style_cell(ws, row, 5, ov, ft=font(DARK_GREEN, 9), fl=fill(LIGHT_GREEN), align=ALC, border=thin_border)
        style_cell(ws, row, 6, '', fl=fill(MED_GRAY), border=thin_border)
        style_cell(ws, row, 7, cv, ft=font(AMBER, 9), fl=fill(LIGHT_AMBER), align=ALC, border=thin_border)
        for cc in [8, 9, 10]:
            style_cell(ws, row, cc, '', fl=fill(bg), border=thin_border)
        row += 1

    # Footer
    row += 1
    merge_and_style(ws, row, 1, 10,
        '⚠ The scenario selector (B6) drives the entire model. Editable inputs are in the Scenarios sheet. '
        'Company settings are in the Company Info sheet.',
        ft=font(AMBER, 9, italic=True), fl=fill(LIGHT_AMBER), align=AL)

    # Freeze panes
    ws.freeze_panes = 'B10'
    return ws
