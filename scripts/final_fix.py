"""
Final fix: Ensure calcProperties force full recalculation on load,
style the control cell prominently, and verify everything end-to-end.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, sys, shutil

sys.stdout.reconfigure(encoding='utf-8')

SRC = os.path.join(os.environ['USERPROFILE'], 'Downloads', 'WOLF_Financial_Model_Fixed.xlsx')
OUT = os.path.join(os.environ['USERPROFILE'], 'Downloads', 'WOLF_Financial_Model_Final.xlsx')

shutil.copy2(SRC, OUT)

wb = openpyxl.load_workbook(OUT)

# 1. Ensure calc properties force full recalculation
print("=== Setting Calculation Properties ===")
if wb.calculation is None:
    wb.calculation = openpyxl.workbook.properties.CalcProperties()
wb.calculation.calcMode = 'auto'
wb.calculation.fullCalcOnLoad = True
wb.calculation.iterate = True
wb.calculation.iterateCount = 100
wb.calculation.iterateDelta = 0.001
print(f"  calcMode: {wb.calculation.calcMode}")
print(f"  fullCalcOnLoad: {wb.calculation.fullCalcOnLoad}")
print(f"  iterate: {wb.calculation.iterate}")

# 2. Style the control cell prominently
print("\n=== Styling Control Cell B64 ===")
asheet = wb['Assumptions']
ctrl = asheet['B64']
print(f"  Current value: {repr(ctrl.value)}")

# Amber background, thick amber border, bold navy font
amber_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
amber_border = Border(
    left=Side(style='thick', color='FFA500'),
    right=Side(style='thick', color='FFA500'),
    top=Side(style='thick', color='FFA500'),
    bottom=Side(style='thick', color='FFA500'),
)
navy_font = Font(name='Calibri', size=12, bold=True, color='000080')

ctrl.fill = amber_fill
ctrl.border = amber_border
ctrl.font = navy_font
ctrl.alignment = Alignment(horizontal='center', vertical='center')

# Label cell A64
label = asheet['A64']
label.value = 'Active Scenario'
label.fill = amber_fill
label.border = amber_border
label.font = navy_font
label.alignment = Alignment(horizontal='right', vertical='center')

# 3. Style IF formula cells with dark green
print("\n=== Styling IF Formula Cells ===")
green_font = Font(name='Calibri', size=10, color='004400')
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

KEY_ROWS = [5, 6, 7, 8, 9, 10, 22, 26, 32]
styled_count = 0
for r in KEY_ROWS:
    for c in range(5, 10):
        cell = asheet.cell(r, c)
        val = cell.value
        if val and isinstance(val, str) and 'IF' in val:
            cell.font = green_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='right')
            styled_count += 1

print(f"  Styled {styled_count} IF formula cells with dark green font")

# 4. Verify sheet order (Dashboard should be first)
print("\n=== Sheet Order ===")
for i, name in enumerate(wb.sheetnames):
    print(f"  {i}: {name}")

# 5. Save
wb.save(OUT)
print(f"\n=== Saved: {OUT} ===")

# 6. Re-read and verify
wb2 = openpyxl.load_workbook(OUT)
print("\n=== Final Verification ===")
print(f"  calcMode: {wb2.calculation.calcMode if wb2.calculation else 'N/A'}")
print(f"  fullCalcOnLoad: {wb2.calculation.fullCalcOnLoad if wb2.calculation else 'N/A'}")

a2 = wb2['Assumptions']
print(f"  B64 value: {repr(a2['B64'].value)}")
print(f"  E5 formula: {repr(a2['E5'].value)}")

d2 = wb2['Dashboard']
print(f"  Dashboard B6: {repr(d2['B6'].value)}")

# Verify data validations still present
for dv in d2.data_validations.dataValidation:
    for cr in dv.cells.ranges:
        if 'B6' in str(cr):
            print(f"  B6 validation: type={dv.type}, formula1={dv.formula1}")

# Count scenario blocks
s2 = wb2['Scenarios']
block_count = 0
for r in range(1, s2.max_row + 1):
    v = str(s2.cell(r, 1).value or '')
    if '\u258e' in v:
        block_count += 1
        safe = v.replace('\u258e', '|')
        print(f"  Block {block_count}: {safe}")

# Count IF formulas
if_count = 0
for r in KEY_ROWS:
    for c in range(5, 10):
        v = a2.cell(r, c).value
        if v and isinstance(v, str) and 'IF' in v:
            if_count += 1

print(f"\n  Scenario blocks: {block_count}/3")
print(f"  IF formulas: {if_count}/45")
print(f"  Control cell: {repr(a2['B64'].value)}")

wb2.close()

if block_count == 3 and if_count == 45:
    print("\n  *** ALL CHECKS PASSED - File is ready! ***")
else:
    print("\n  *** SOME CHECKS FAILED ***")

print(f"\n  Output: {OUT}")
