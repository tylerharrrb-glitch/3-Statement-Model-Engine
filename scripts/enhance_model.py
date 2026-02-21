"""
Main orchestrator: enhances the financial model Excel workbook.
Run: python scripts/enhance_model.py
"""
import sys, os, shutil
sys.path.insert(0, os.path.dirname(__file__))
import openpyxl
from openpyxl.utils import get_column_letter
from enhance_helpers import *
from build_company_info import build_company_info
from build_scenarios import build_scenarios
from build_dashboard import build_dashboard
from openpyxl.styles import Alignment

# ── Paths ─────────────────────────────────────────────
SRC = r"C:\Users\user\Downloads\WOLF_Financial_Model.xlsx"
WORK = r"C:\Users\user\Desktop\3_Statement_Model_Engine\model_enhanced.xlsx"
OUT = r"C:\Users\user\Downloads\Demo_Company_Inc__Financial_Model_v3.xlsx"

print(f"Source: {SRC} ({os.path.getsize(SRC)} bytes)")
shutil.copy2(SRC, WORK)
print(f"Copied to working file: {WORK}")

wb = openpyxl.load_workbook(WORK)
print(f"Loaded. Sheets: {wb.sheetnames}")

# ══════════════════════════════════════════════════════
# FEATURE 1: Company Info
# ══════════════════════════════════════════════════════
print("\n[1/6] Building Company Info sheet...")
build_company_info(wb)

# ══════════════════════════════════════════════════════
# FEATURE 2: Scenarios
# ══════════════════════════════════════════════════════
print("[2/6] Building Scenarios sheet...")
build_scenarios(wb)

# ══════════════════════════════════════════════════════
# FEATURE 3: Dashboard
# ══════════════════════════════════════════════════════
print("[3/6] Building Dashboard sheet...")
build_dashboard(wb)

# ══════════════════════════════════════════════════════
# WIRING: Scenario Selector → Assumptions
# ══════════════════════════════════════════════════════
print("[4/6] Wiring scenario selector into Assumptions...")
ws_a = wb['Assumptions']
AL = Alignment(vertical='center', wrap_text=True)
ALC = Alignment(vertical='center', horizontal='center')

# Add scenario control block at rows 62-67 (safely below existing data at row 61)
r = 62
merge_and_style(ws_a, r, 1, 9, '── SCENARIO CONTROL ──',
                ft=font(WHITE, 11, True), fl=fill(MED_BLUE), align=ALC)
r = 63
style_cell(ws_a, r, 1, 'Active Scenario', ft=font(NAVY, 10, True), fl=fill(LIGHT_GRAY), align=AL, border=thin_border)
merge_and_style(ws_a, r, 2, 9, "='Dashboard'!B6",
                ft=font(AMBER, 12, True), fl=fill(LIGHT_AMBER), align=ALC)
for c in range(2, 10):
    ws_a.cell(r, c).border = thick_amber_border

r = 64
merge_and_style(ws_a, r, 1, 9,
    '=IF($B$63="Base Case","Central planning — stable market conditions",'
    'IF($B$63="Optimistic","Bull case — strong execution","Bear case — headwinds"))',
    ft=font(DARK_GRAY, 9, italic=True), fl=fill(WHITE), align=AL)

r = 65
merge_and_style(ws_a, r, 1, 9,
    '⚠ Change the scenario ONLY via the Dashboard dropdown (Dashboard!B6). Do not edit this cell directly.',
    ft=font(AMBER, 9, italic=True), fl=fill(LIGHT_AMBER), align=AL)

r = 66
merge_and_style(ws_a, r, 1, 9,
    'Scenario-driven rows: Rev Growth (5), COGS (6), SG&A (7), R&D (8), Other OpEx (9), Tax (10), CapEx (22), Int Rate (26), Div Payout (32)',
    ft=font(DARK_GRAY, 9), fl=fill(LIGHT_GRAY), align=AL)

r = 67
merge_and_style(ws_a, r, 1, 9,
    'The Scenarios sheet is the single source of truth. Edit scenario values there, not here.',
    ft=font(DARK_GRAY, 9), fl=fill(LIGHT_GRAY), align=AL)

# ── Replace projection-period cells with IF formulas ──
# Scenarios sheet layout: blocks start at rows 4(+2=data), 17(+2=data), 30(+2=data)
# Actually from the build_scenarios code, the blocks use _write_block which starts:
# Block 1 starts at row 4: title=4, header=5, data rows=6-14
# Block 2 starts at row 15+1=16: title=16, header=17, data rows=18-26  (after +1 spacer) 
# Block 3 starts at row 27+1=28: title=28, header=29, data rows=30-38

# Let me recalculate: _write_block returns the next row after 2+9=11 rows
# Block 1: start=4, title=4, header=5, data=6..14, returns 15
# spacer at 15, so block2 starts at 16
# Block 2: start=16, title=16, header=17, data=18..26, returns 27
# spacer at 27, so block3 starts at 28
# Block 3: start=28, title=28, header=29, data=30..38

# Scenario data rows offset within each block: 0=RevGrowth, 1=COGS, etc.
SCEN_BASE_ROWS = [6, 7, 8, 9, 10, 11, 12, 13, 14]     # data rows for Base
SCEN_OPT_ROWS =  [18, 19, 20, 21, 22, 23, 24, 25, 26]  # data rows for Optimistic
SCEN_CONS_ROWS = [30, 31, 32, 33, 34, 35, 36, 37, 38]  # data rows for Conservative

# Assumptions projection col E-I → Scenarios col B-F
# assump_col: 5(E) 6(F) 7(G) 8(H) 9(I)
# scen_col:   B    C    D    E    F

for assump_idx, key in enumerate(ASSUMP_ORDER):
    assump_row = ASSUMP_ROWS[key]
    base_scen_row = SCEN_BASE_ROWS[assump_idx]
    opt_scen_row = SCEN_OPT_ROWS[assump_idx]
    cons_scen_row = SCEN_CONS_ROWS[assump_idx]

    for assump_col in range(5, 10):  # E=5 through I=9
        scen_col_letter = get_column_letter(assump_col - 3)  # E(5)→B(2), F(6)→C(3), etc.
        
        formula = (
            f'=IF($B$63="Base Case",Scenarios!{scen_col_letter}{base_scen_row},'
            f'IF($B$63="Optimistic",Scenarios!{scen_col_letter}{opt_scen_row},'
            f'Scenarios!{scen_col_letter}{cons_scen_row}))'
        )
        cell = ws_a.cell(row=assump_row, column=assump_col, value=formula)
        cell.font = font(DARK_GREEN, 10)
        cell.border = thin_border
        cell.number_format = '0.0%'

print(f"  Wired {len(ASSUMP_ORDER)} assumption rows × 5 projection columns = {len(ASSUMP_ORDER)*5} formulas")

# ══════════════════════════════════════════════════════
# SHEET ORDER & TAB COLORS
# ══════════════════════════════════════════════════════
print("[5/6] Reordering sheets and applying tab colors...")

DESIRED_ORDER = [
    'Dashboard', 'Company Info', 'Scenarios', 'Assumptions',
    'Income Statement', 'Balance Sheet', 'Cash Flow Statement', 'Ratios',
    'Working Capital', 'Depreciation Schedule', 'Debt Schedule', 'CBE Banking Metrics',
]

# Delete old Summary Dashboard if exists
if 'Summary Dashboard' in wb.sheetnames:
    del wb['Summary Dashboard']

# Reorder
current = list(wb.sheetnames)
order_map = []
for name in DESIRED_ORDER:
    if name in current:
        order_map.append(current.index(name))
wb.move_sheet('Dashboard', offset=-wb.sheetnames.index('Dashboard'))
# Use a simpler approach: move each sheet to its target position
for target_idx, name in enumerate(DESIRED_ORDER):
    if name in wb.sheetnames:
        current_idx = wb.sheetnames.index(name)
        offset = target_idx - current_idx
        if offset != 0:
            wb.move_sheet(name, offset=offset)

TAB_COLORS = {
    'Dashboard': NAVY,
    'Company Info': MED_BLUE,
    'Scenarios': GREEN,
    'Assumptions': GRAY_TAB,
    'Income Statement': STD_BLUE, 'Balance Sheet': STD_BLUE,
    'Cash Flow Statement': STD_BLUE, 'Ratios': STD_BLUE,
    'Working Capital': BROWN, 'Depreciation Schedule': BROWN, 'Debt Schedule': BROWN,
    'CBE Banking Metrics': RED,
}
for name, color in TAB_COLORS.items():
    if name in wb.sheetnames:
        wb[name].sheet_properties.tabColor = color

# ── Update A1 on statement sheets with Company Info reference ──
title_sheets = {
    'Income Statement': 'Income Statement',
    'Balance Sheet': 'Balance Sheet', 
    'Cash Flow Statement': 'Cash Flow Statement',
    'Ratios': 'Financial Ratios',
    'Working Capital': 'Working Capital Schedule',
    'Depreciation Schedule': 'PP&E Rollforward',
    'Debt Schedule': 'Debt Schedule',
}
for sheet_name, subtitle in title_sheets.items():
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws['A1'] = f"=\"'\" & 'Company Info'!B4 & \"' — {subtitle}\""

# ══════════════════════════════════════════════════════
# SAVE & DELIVER
# ══════════════════════════════════════════════════════
print("[6/6] Saving...")
wb.save(WORK)
print(f"Saved to: {WORK} ({os.path.getsize(WORK)} bytes)")

# Copy to output
shutil.copy2(WORK, OUT)
print(f"Copied to: {OUT}")

# ── Verify ────────────────────────────────────────────
wb2 = openpyxl.load_workbook(WORK, data_only=False)
total_formulas = 0
for name in wb2.sheetnames:
    ws = wb2[name]
    count = 0
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                count += 1
    total_formulas += count
    print(f"  {name}: {count} formulas")
wb2.close()

print(f"\n{'='*60}")
print(f"✓ COMPLETE")
print(f"  Total formulas: {total_formulas}")
print(f"  Sheet order: {wb.sheetnames}")
print(f"  Active scenario: Base Case (Dashboard B6)")
print(f"  Scenario selector cascades to all model formulas via Assumptions")  
print(f"  Company Name (Company Info B4) → Dashboard + all sheet titles")
print(f"  Scenarios sheet = single source of truth for assumptions")
print(f"{'='*60}")
