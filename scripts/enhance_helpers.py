"""Shared helpers for Excel model enhancement."""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ── Colors ────────────────────────────────────────────
NAVY = '1F3864'
MED_BLUE = '2E75B6'
STD_BLUE = '4472C4'
LIGHT_BLUE = 'D6E4F0'
VERY_LIGHT_BLUE = 'EBF3FB'
WHITE = 'FFFFFF'
LIGHT_GRAY = 'F2F2F2'
MED_GRAY = 'D9D9D9'
DARK_GRAY = '404040'
INPUT_BLUE = '0000CC'
SCENARIO_BLUE = '000080'
GREEN = '1A7A4A'
LIGHT_GREEN = 'E8F4EA'
DARK_GREEN = '004400'
AMBER = 'B7791F'
LIGHT_AMBER = 'FFF3E0'
BROWN = '8B4000'
RED = 'C0392B'
GRAY_TAB = '7F7F7F'
BORDER_COLOR = 'CCCCCC'

# ── Reusable styles ──────────────────────────────────
thin_border = Border(
    left=Side(style='thin', color=BORDER_COLOR),
    right=Side(style='thin', color=BORDER_COLOR),
    top=Side(style='thin', color=BORDER_COLOR),
    bottom=Side(style='thin', color=BORDER_COLOR),
)
thick_blue_border = Border(
    left=Side(style='medium', color=MED_BLUE),
    right=Side(style='medium', color=MED_BLUE),
    top=Side(style='medium', color=MED_BLUE),
    bottom=Side(style='medium', color=MED_BLUE),
)
thick_amber_border = Border(
    left=Side(style='medium', color=AMBER),
    right=Side(style='medium', color=AMBER),
    top=Side(style='medium', color=AMBER),
    bottom=Side(style='medium', color=AMBER),
)

def fill(color): return PatternFill('solid', fgColor=color)
def font(color=DARK_GRAY, size=10, bold=False, italic=False):
    return Font(name='Calibri', size=size, bold=bold, italic=italic, color=color)

def style_cell(ws, row, col, value, ft=None, fl=None, align=None, border=None, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    if ft: c.font = ft
    if fl: c.fill = fl
    if align: c.alignment = align
    if border: c.border = border or thin_border
    if fmt: c.number_format = fmt
    return c

def merge_and_style(ws, r, c1, c2, value, ft=None, fl=None, align=None):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    c = ws.cell(row=r, column=c1, value=value)
    if ft: c.font = ft
    if fl: c.fill = fl
    if align: c.alignment = align
    for col in range(c1, c2+1):
        ws.cell(row=r, column=col).border = thin_border
        if fl: ws.cell(row=r, column=col).fill = fl
    return c

def apply_row_border(ws, row, max_col):
    for c in range(1, max_col+1):
        ws.cell(row=row, column=c).border = thin_border

# ── Assumption row → Scenario row mapping ────────────
# Assumptions rows that get scenario-driven formulas
ASSUMP_ROWS = {
    'rev_growth': 5, 'cogs': 6, 'sga': 7, 'rd': 8, 'other_opex': 9,
    'tax': 10, 'capex': 22, 'int_rate': 26, 'div_payout': 32,
}
# Order for scenario blocks
ASSUMP_ORDER = ['rev_growth', 'cogs', 'sga', 'rd', 'other_opex', 'tax', 'capex', 'int_rate', 'div_payout']
ASSUMP_LABELS = [
    'Revenue Growth Rate', 'COGS % of Revenue', 'SG&A % of Revenue',
    'R&D % of Revenue', 'Other OpEx % of Revenue', 'Tax Rate',
    'CapEx % of Revenue', 'Interest Rate (on Debt)', 'Dividend Payout Ratio',
]
# Scenario data block start rows (in Scenarios sheet)
BASE_START = 6; OPT_START = 19; CONS_START = 32
# Projection cols: Assumptions E-I → Scenarios B-F
PROJ_COL_MAP = {5: 2, 6: 3, 7: 4, 8: 5, 9: 6}  # assump_col → scen_col
