"""
PHASE 1 - Full diagnosis of WOLF Excel scenario wiring.
"""
import openpyxl
from openpyxl.utils import get_column_letter
import os, shutil, sys

# Force UTF-8 stdout
sys.stdout.reconfigure(encoding='utf-8')

SRC = os.path.join(os.environ['USERPROFILE'], 'Downloads', 'WOLF_Financial_Model_Fixed.xlsx')
WORK = os.path.join(os.environ['USERPROFILE'], 'Downloads', 'WOLF_working.xlsx')

shutil.copy2(SRC, WORK)
print(f"Working copy: {WORK}\n")

# === 1a. Map Scenarios sheet ===
print("=" * 70)
print("1a. SCENARIOS SHEET - Complete Map")
print("=" * 70)
wb_data = openpyxl.load_workbook(SRC, data_only=True)
scen = wb_data['Scenarios']
print(f"  Dimensions: {scen.dimensions}")
print(f"  Max row: {scen.max_row}, Max col: {scen.max_column}")

block_starts = []
data_rows_by_block = {}
current_block = None

for r in range(1, scen.max_row + 1):
    row_vals = []
    for c in range(1, scen.max_column + 1):
        v = scen.cell(r, c).value
        if v is not None:
            row_vals.append((get_column_letter(c), v))
    if not row_vals:
        continue
    
    first_val = str(row_vals[0][1])
    
    # Block title marker
    if '\u258e' in first_val or 'BLOCK' in first_val.upper() or (first_val.startswith('|') and 'CASE' in first_val.upper()):
        block_starts.append(r)
        block_name = first_val.replace('\u258e', '|').strip()
        current_block = block_name
        data_rows_by_block[current_block] = []
        print(f"\n  *** BLOCK TITLE Row {r}: {block_name}")
    elif '---' in first_val or '===' in first_val:
        print(f"  SECTION Row {r}: {first_val[:60]}")
    elif r <= 3:
        safe = first_val.replace('\u2550', '=').replace('\u2014', '-')
        print(f"  Row {r}: {safe[:80]}")
    else:
        label = str(row_vals[0][1])[:38]
        # Collect data values
        nums = []
        for col_letter, v in row_vals[1:]:
            if isinstance(v, (int, float)):
                nums.append(f"{v:.4f}" if abs(v) < 2 else f"{v:,.0f}")
        
        if nums:
            if current_block:
                data_rows_by_block[current_block].append(r)
            print(f"  Row {r}: {label:38s} | {', '.join(nums[:8])}")
        elif 'Assumption' in label or 'Computed' in label:
            print(f"  Row {r}: {label} (header)")

print(f"\n  Total scenario blocks: {len(block_starts)}")
print(f"  Block title rows: {block_starts}")
for bname, rows in data_rows_by_block.items():
    safe = bname.replace('\u258e', '|')
    print(f"  Block '{safe}': {len(rows)} data rows, rows {rows[:5]}...{rows[-3:]}")

# Period columns
print("\n  Period columns:")
for r in range(1, scen.max_row + 1):
    first = scen.cell(r, 1).value
    if first and 'Assumption' in str(first):
        for c in range(1, scen.max_column + 1):
            v = scen.cell(r, c).value
            if v:
                print(f"    Col {get_column_letter(c)} = {v}")
        break

wb_data.close()

# === 1b. Dashboard B6 ===
print("\n" + "=" * 70)
print("1b. DASHBOARD B6 - Scenario selector")
print("=" * 70)
wb_form = openpyxl.load_workbook(SRC)
dash = wb_form['Dashboard']
b6 = dash['B6']
print(f"  B6 value: {repr(b6.value)}")
print(f"  B6 data_type: {b6.data_type}")

for dv in dash.data_validations.dataValidation:
    for cell_range in dv.cells.ranges:
        if 'B6' in str(cell_range):
            print(f"  DataValidation on {cell_range}:")
            print(f"    type: {dv.type}")
            print(f"    formula1: {dv.formula1}")

# === 1c. Assumptions control cell ===
print("\n" + "=" * 70)
print("1c. ASSUMPTIONS - Control cell")
print("=" * 70)
asheet = wb_form['Assumptions']
control_addr = None
for r in range(1, asheet.max_row + 1):
    for c in range(1, min(asheet.max_column + 1, 15)):
        cell = asheet.cell(r, c)
        val = cell.value
        if val and 'Dashboard' in str(val):
            addr = f"{get_column_letter(c)}{r}"
            print(f"  Found at {addr}: {repr(val)}")
            control_addr = addr
            break
    if control_addr:
        break

if not control_addr:
    print("  No Dashboard reference found! Searching for 'Active Scenario'...")
    for r in range(1, asheet.max_row + 1):
        v = asheet.cell(r, 1).value
        if v and 'Active' in str(v):
            print(f"    A{r}: {v}")
            b_val = asheet.cell(r, 2).value
            print(f"    B{r}: {repr(b_val)}")
            control_addr = f"B{r}"

# === 1d. 9 key assumption rows ===
print("\n" + "=" * 70)
print("1d. ASSUMPTIONS - 9 key rows, projection columns")
print("=" * 70)

# First, find what columns are projection columns
print("  Period headers:")
for c in range(1, 15):
    v = asheet.cell(1, c).value
    if v:
        print(f"    {get_column_letter(c)}1 = {v}")

TARGET_LABELS = [
    'Revenue Growth Rate',
    'COGS % of Revenue', 
    'SG&A % of Revenue',
    'R&D % of Revenue',
    'Other OpEx % of Revenue',
    'Tax Rate',
    'CapEx % of Revenue',
    'Interest Rate',
    'Dividend Payout',
]

found_rows = {}
for r in range(1, asheet.max_row + 1):
    label = str(asheet.cell(r, 1).value or '')
    for target in TARGET_LABELS:
        if target.lower() in label.lower() and target not in found_rows:
            found_rows[target] = r
            break

# Find first projection column
proj_start_col = None
for c in range(2, 15):
    v = str(asheet.cell(1, c).value or '')
    if 'E' in v and any(y in v for y in ['2024', '2025', '2026']):
        proj_start_col = c
        break

if proj_start_col:
    print(f"\n  First projection column: {get_column_letter(proj_start_col)} (col {proj_start_col})")
else:
    print("\n  WARNING: Could not determine first projection column")
    proj_start_col = 5  # default to E

if_formula_count = 0
hardcoded_count = 0

for target in TARGET_LABELS:
    r = found_rows.get(target)
    if r is None:
        print(f"\n  NOT FOUND: {target}")
        continue
    
    label = asheet.cell(r, 1).value
    print(f"\n  Row {r}: {label}")
    for c in range(proj_start_col, proj_start_col + 5):
        cell = asheet.cell(r, c)
        col = get_column_letter(c)
        val = cell.value
        if val and isinstance(val, str) and ('IF' in val.upper() or val.startswith('=')):
            print(f"    {col}{r}: FORMULA -> {val[:120]}")
            if 'IF' in val.upper():
                if_formula_count += 1
        elif isinstance(val, (int, float)):
            if abs(val) < 2:
                print(f"    {col}{r}: NUMBER -> {val:.4f} ({val*100:.1f}%)")
            else:
                print(f"    {col}{r}: NUMBER -> {val:,.2f}")
            hardcoded_count += 1
        else:
            print(f"    {col}{r}: {type(val).__name__} -> {repr(val)}")

# === 1e. Cascade path ===
print("\n" + "=" * 70)
print("1e. CASCADE PATH")
print("=" * 70)

is_sheet = wb_form['Income Statement']
print(f"  Income Statement E2: {repr(is_sheet['E2'].value)}")
print(f"  Income Statement F2: {repr(is_sheet['F2'].value)}")
print(f"  Assumptions E3: {repr(asheet['E3'].value)}")
print(f"  Assumptions E4: {repr(asheet['E4'].value)}")
print(f"  Assumptions E5: {repr(asheet['E5'].value)}")

# === Summary ===
print("\n" + "=" * 70)
print("DIAGNOSIS SUMMARY")
print("=" * 70)
print(f"  Scenario blocks: {len(block_starts)} (need 3)")
print(f"  IF formulas in key rows: {if_formula_count} (need 45)")
print(f"  Hardcoded numbers in key rows: {hardcoded_count}")
print(f"  Control cell address: {control_addr}")
print(f"  Dashboard B6: {repr(dash['B6'].value)}")

if len(block_starts) < 3:
    print(f"\n  BROKEN: Only {len(block_starts)} scenario block(s)")
if if_formula_count == 0:
    print(f"  BROKEN: No IF formulas - all projection cells are hardcoded")
if if_formula_count > 0 and if_formula_count < 45:
    print(f"  PARTIAL: Only {if_formula_count}/45 IF formulas written")

wb_form.close()
print("\n=== Phase 1 Complete ===")
